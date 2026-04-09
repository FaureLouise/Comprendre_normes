import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import zipfile
import io
import openpyxl
from pandas import ExcelWriter
from streamlit_sortables import sort_items
from scipy.stats import norm
from matplotlib.patches import FancyBboxPatch
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook

# Charger le fichier Excel
file_path = 'NORMES_FEV_25.xlsx'
excel_data = pd.ExcelFile(file_path)

# Liste des groupes d'âge (onglets du fichier)
age_groups = excel_data.sheet_names

if "age_selected" not in st.session_state:
    st.session_state["age_selected"] = False

if "scores_entered" not in st.session_state:
    st.session_state["scores_entered"] = False

if "age_data" not in st.session_state:
    st.session_state["age_data"] = pd.DataFrame()

if "missing_norms" not in st.session_state:
    st.session_state["missing_norms"] = []

st.markdown(
    """
    <div style="text-align: center; font-size: 40px; font-weight: bold;">
        Batterie COMPRENDRE
    </div>
    """,
    unsafe_allow_html=True
)

#Âge ET ID
st.header("Étape 1 : Sélectionnez le groupe d'âge")
selected_age_group = st.selectbox("Sélectionnez le groupe d'âge de l'enfant :", age_groups)
child_id = st.text_input("Saisissez l'ID de l'enfant :", value="", placeholder="ID de l'enfant")

if st.button("Passer à l'étape suivante"):
    if not child_id.strip(): 
        st.error("Veuillez saisir un ID valide avant de continuer.")
    else:
        st.session_state["age_selected"] = True
        st.session_state["child_id"] = child_id 
        st.success(f"ID {child_id} et âge {selected_age_group} confirmés.")
        
        
def load_age_data(sheet_name, excel_file):
    try:
        return pd.read_excel(excel_file, sheet_name=sheet_name, engine="openpyxl")
    except Exception as e:
        st.error(f"Erreur lors du chargement des données : {e}")
        return pd.DataFrame()


if st.session_state["age_selected"]:
    st.header("Étape 2 : Entrez les scores")
    age_data = load_age_data(selected_age_group, excel_data)

    if age_data.empty:
        st.error("Impossible de charger les données pour le groupe d'âge sélectionné.")
    else:
        age_data = age_data[["Tâche", "Moyenne", "Ecart-type", "Minimum", 
                             "5e percentile", "10e percentile", "Q1", 
                             "Q2 - mediane", "Q3", "90e percentile", "Maximum"]].dropna()

        # Liste des catégories avec les tâches regroupées par paires
        categories = {
            "Langage": [
                ("Discrimination Phonologique", "Décision Lexicale Auditive"),
                ("Mots Outils", "Stock Lexical"),
                ("Compréhension Syntaxique", "")
            ],
            "Mémoire de Travail Verbale": [
                ("Mémoire de travail verbale endroit empan", "Mémoire de travail verbale endroit brut"),
                ("Mémoire de travail verbale envers empan", "Mémoire de travail verbale envers brut")
            ],
            "Mémoire de Travail Non Verbale": [
                ("Mémoire de travail non verbale endroit empan", "Mémoire de travail non verbale endroit brut"),
                ("Mémoire de travail non verbale envers empan", "Mémoire de travail non verbale envers brut")  
            ],   
            "Inhibition verbale": [
                ("Inhibition verbale incongruent score", "Inhibition verbale incongruent temps")
            ],
            "Inhibition non verbale": [
                ("Inhibition non verbale incongruent score", "Inhibition non verbale incongruent temps")
            ]
        }

        # Collecte des scores utilisateur et calculs d'interférences
        user_scores = []
        inhibition_scores = {}
        missing_norms = []

        for category, task_pairs in categories.items():
            st.subheader(category)
            for task1, task2 in task_pairs:
                col1, col2 = st.columns(2)

                # Colonne 1 : Saisie pour task1
                with col1:
                    if task1 and task1 in age_data["Tâche"].values:
                        score1 = st.text_input(f"{task1} :", value="")
                        if score1.strip(): 
                            try:
                                score1 = float(score1)
                                user_scores.append({"Tâche": task1, "Score Enfant": score1})
                                inhibition_scores[task1] = score1
                            except ValueError:
                                st.error(f"Valeur non valide pour {task1}. Veuillez entrer un nombre.")
                                inhibition_scores[task1] = score1
                    elif task1:
                        st.warning(f"Pas de normes disponibles pour {task1}")
                        missing_norms.append(task1)

                # Colonne 2 : Saisie pour task2
                with col2:
                    if task2 and task2 in age_data["Tâche"].values:
                        score2 = st.text_input(f"{task2} :", value="")
                        if score2.strip():  
                            try:
                                score2 = float(score2)
                                user_scores.append({"Tâche": task2, "Score Enfant": score2})
                                inhibition_scores[task2] = score2
                            except ValueError:
                                st.error(f"Valeur non valide pour {task2}. Veuillez entrer un nombre.")
                                inhibition_scores[task2] = score2
                    elif task2:
                        st.warning(f"Pas de normes disponibles pour {task2}")
                        missing_norms.append(task2)

        scores_df = pd.DataFrame(user_scores, columns=["Tâche", "Score Enfant"])

        # Inverser les Z-scores pour les variables de temps d'inhibition
        time_variables = [
            "Inhibition verbale incongruent temps",
            "Inhibition non verbale incongruent temps"]
      
        
        # Fusionner avec les données originales pour les calculs
        merged_data = pd.merge(age_data, scores_df, on="Tâche", how="left")
        merged_data["Z-Score"] = (merged_data["Score Enfant"] - merged_data["Moyenne"]) / merged_data["Ecart-type"]
        merged_data.loc[merged_data["Tâche"].isin(time_variables), "Z-Score"] *= -1

        merged_data["Z-Score"] = pd.to_numeric(merged_data["Z-Score"], errors="coerce")
        merged_data = merged_data.dropna(subset=["Z-Score"])

        merged_data["Percentile (%)"] = norm.cdf(merged_data["Z-Score"]) * 100

        filled_data = merged_data[~merged_data["Score Enfant"].isna()]
        filled_data = filled_data.drop_duplicates(subset="Tâche")

        # Bouton pour confirmer les scores
        if st.button("Confirmer les scores et afficher les résultats"):
            st.session_state["scores_entered"] = True
            st.session_state["age_data"] = filled_data
            st.session_state["missing_norms"] = missing_norms

# Étape 3 : Résultats
    categories_mapping = {
        "Langage": [
            "Discrimination Phonologique", "Décision Lexicale Auditive",
            "Mots Outils", "Stock Lexical", "Compréhension Syntaxique"
        ],
        "Mémoire de Travail": [
            "Mémoire de travail verbale endroit empan", "Mémoire de travail verbale endroit brut",
            "Mémoire de travail verbale envers empan", "Mémoire de travail verbale envers brut",
            "Mémoire de travail non verbale endroit empan", "Mémoire de travail non verbale endroit brut",
            "Mémoire de travail non verbale envers empan", "Mémoire de travail non verbale envers brut"
        ],
        "Inhibition": [
            "Inhibition verbale incongruent score",
            "Inhibition verbale incongruent temps",
            "Inhibition non verbale incongruent score",
            "Inhibition non verbale incongruent temps"
        ]
    }

    task_name_mapping = {
        "Discrimination Phonologique": "Discrimination\nPhonologique",
        "Décision Lexicale Auditive": "Décision\nLexicale\nAuditive",
        "Mots Outils": "Mots\nOutils",
        "Stock Lexical": "Stock\nLexical",
        "Compréhension Syntaxique": "Compréhension\nSyntaxique",
        "Mémoire de travail verbale endroit empan": "Mémoire de travail\nVerbale\nendroit\nempan",
        "Mémoire de travail verbale endroit brut": "Mémoire de travail\nVerbale\nendroit\nbrut",
        "Mémoire de travail verbale envers empan": "Mémoire de travail\nVerbale\nenvers\nempan",
        "Mémoire de travail verbale envers brut": "Mémoire de travail\nVerbale\nenvers\nbrut",
        "Mémoire de travail non verbale endroit empan": "Mémoire de travail\nNon Verbale\nendroit\nempan",
        "Mémoire de travail non verbale endroit brut": "Mémoire de travail\nNon Verbale\nendroit\nbrut",
        "Mémoire de travail non verbale envers empan": "Mémoire de travail\nNon Verbale\nenvers\nempan",
        "Mémoire de travail non verbale envers brut": "Mémoire de travail\nNon Verbale\nenvers\nbrut",
        "Inhibition verbale incongruent score": "Inhibition\nVerbale\nIncongruent\nscore",
        "Inhibition verbale incongruent temps": "Inhibition\nVerbale\nIncongruent\ntemps",
        "Inhibition non verbale incongruent score": "Inhibition\nNon Verbale\nIncongruent\nscore",
        "Inhibition non verbale incongruent temps": "Inhibition\nNon Verbale\nIncongruent\ntemps"
    }

    # Ajouter la colonne "Catégorie" pour chaque tâche
    def plot_grouped_scores(data, selected_tasks):
        category_colors = {
            "Langage": "#3798da",
            "Mémoire de Travail": "#eca113",
            "Inhibition": "#8353da",
            "Autre": "gray"
        }

        # Filtrer les données pour inclure uniquement les tâches sélectionnées
        data = data[data["Tâche"].isin(selected_tasks)]

        # Liste des tâches (abrégées) et leurs Z-scores
        tasks = data["Tâche"].map(task_name_mapping).tolist()
        percentiles = data["Percentile (%)"].tolist()

        positions = np.arange(len(tasks))  

        # Ajouter une colonne pour les positions dans le DataFrame
        data["Position"] = positions

        # Créer la figure
        fig_width = 14
        fig_height = max(10, len(tasks) * 1.5)
        fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=300)

        # Tracer les points pour chaque tâche
        point_colors = data["Catégorie"].map(category_colors)
        ax.scatter(percentiles, positions, color=point_colors, s=100, zorder=3)

        # Ajouter les scores de l'enfant avec un cadre coloré autour
        for i, (score, category, task_name, mean, std_dev) in enumerate(zip(data["Score Enfant"], data["Catégorie"], data["Tâche"], data["Moyenne"], data["Ecart-type"])):
            color = category_colors.get(category, "gray")  

            # Calculer la hauteur en fonction de l'espacement des points sur l'axe Y
            if len(positions) > 1:  
                spacing = positions[1] - positions[0]  
            else:
                spacing = 1  

            box_height = spacing * 0.2  # Ajuster la hauteur proportionnellement à l'espacement
            vertical_offset = box_height / 2  # Centrer le cadre autour du point

            # Ajouter le cadre
            bbox = FancyBboxPatch(
                (105, positions[i] - vertical_offset),  # Coordonnées (x, y) centrées
                width=28,  # Largeur du cadre
                height=box_height,  # Hauteur ajustée dynamiquement
                boxstyle="square,pad=0.1",  # Angles arrondis avec padding
                linewidth=3,  # Épaisseur de la bordure
                edgecolor=color,  # Couleur de la bordure
                facecolor="white",  # Couleur de fond
                zorder=1  # Couche d'affichage
            )
            ax.add_patch(bbox)  # Ajouter le cadre au graphique
            
            # Formatage du texte avec le score en gras
            score_text = f"$\\bf{{{score:.0f}}}$\n[M = {mean:.1f} ± {std_dev:.1f}]"

            # Ajouter le texte centré dans le cadre
            ax.text(
                x=119,  # Position X centrée dans le cadre
                y=positions[i],  # Position Y alignée verticalement au centre
                s=score_text,  # Texte formaté
                fontsize=13,
                color="black",  # Couleur du texte
                ha="center",  # Alignement horizontal centré
                va="center",  # Alignement vertical centré
                zorder=2,  # Couche d'affichage au-dessus du cadre
                usetex=False  # Utilisation de Matplotlib sans dépendance à LaTeX
            )

        # Ajouter des zones colorées pour les catégories
        ax.fill_betweenx(range(-1, len(tasks)+1), 0, 3, color="#d44646", alpha=0.2, zorder=1)  # Zone rouge
        ax.fill_betweenx(range(-1, len(tasks)+1), 3, 15, color="#f5a72f", alpha=0.2, zorder=1)  # Zone orange
        ax.fill_betweenx(range(-1, len(tasks)+1), 15, 85, color="#60cd72", alpha=0.2, zorder=1)  # Zone verte
        ax.fill_betweenx(range(-1, len(tasks)+1), 85, 97, color="#8ddf9b", alpha=0.2, zorder=1)  # Zone vert clair
        ax.fill_betweenx(range(-1, len(tasks)+1), 97, 100, color="#aedeb6", alpha=0.2, zorder=1)  # Zone bleue


        # Ligne de référence Z=0
        ax.axvline(50, color="black", linestyle="--", linewidth=0.8, zorder=2)
        
        ax.set_xlim(0, 140)  # Axe X : percentiles de 0 à 100
        ax.set_ylim(-1, len(tasks))

        # Configurer les ticks et les labels
        ax.set_xticks([0, 3, 15, 50, 85, 97, 100])
        ax.set_xticklabels(["0", "3", "15", "50", "85", "97", "100"], fontsize=11, fontweight="bold", rotation = -40)
        ax.set_yticks(positions)
        ax.set_yticklabels(tasks, fontsize=16, fontweight="bold")
        ax.set_xlabel("Percentiles (%)", fontsize=14)
        ax.xaxis.set_label_coords(0.85 , -0.02)
        ax.set_ylabel("")

        fig.suptitle(
            "Résultats Batterie Comprendre",
            fontsize=24,
            fontweight="bold",
            x= 0.5, 
            y=1       
        )

        for idx, category in enumerate(category_colors.keys()):
            # Filtrer les données pour cette catégorie
            category_data = data[data["Catégorie"] == category]
            
            # Obtenir les positions et les percentiles pour les tâches dans la catégorie
            category_positions = category_data["Position"].tolist() if not category_data.empty else []
            category_percentiles = category_data["Percentile (%)"].tolist() if not category_data.empty else []

            # Relier les points avec une ligne si la catégorie n'est pas vide
            if category_positions and category_percentiles:
                ax.plot(
                    category_percentiles,  # Les percentiles sur l'axe X
                    category_positions,   # Les positions sur l'axe Y
                    marker="o", linestyle="-", color=category_colors[category],
                    label=category, zorder=4, linewidth=2
                )

        # Ajouter des titres par catégorie sur l'axe Y
        for category, color in category_colors.items():
            # Filtrer les tâches dans la catégorie
            category_data = data[data["Catégorie"] == category]
            
            # Si la catégorie n'est pas vide, ajouter un titre
            if not category_data.empty:
                # Calculer la position moyenne des tâches de la catégorie
                category_positions = category_data["Position"].tolist()
                mid_position = np.mean(category_positions)
                
                # Ajouter le texte pour le titre de la catégorie avec un cadre coloré
                ax.text(
                    x=-40,  # Décalage vers la gauche (en dehors des ticks Y)
                    y=mid_position,
                    s=category.upper(),
                    color="white",  # Couleur du texte
                    fontsize=20,
                    fontweight="bold",
                    ha="center",  # Aligner à droite
                    va="center", 
                    rotation=90,
                    bbox=dict(
                        facecolor=color,  # Couleur de fond
                        edgecolor=color,    # Couleur de la bordure (correspond à la catégorie)
                        boxstyle="round,pad=0.3",  # Bord arrondi avec padding
                        linewidth=2,         # Épaisseur de la bordure
                        alpha=1              # Transparence du fond
                    )
                )


        # Colorer les labels des ticks en fonction des catégories
        for idx, task_label in enumerate(ax.get_yticklabels()):
            if idx < len(data):
                task_category = data.iloc[idx]["Catégorie"]
                task_label.set_color(category_colors.get(task_category, "gray"))
                
        
        for spine in ["top", "right", "bottom", "left"]:
            ax.spines[spine].set_color("white")  # Couleur noire pour la bordure
            ax.spines[spine].set_linewidth(0)    # Épaisseur de la bordure

        # Supprimer la bordure noire à droite
        ax.spines["right"].set_visible(False)


        # Ajuster la mise en page
        plt.subplots_adjust(left=0.3, right=0.95, top=0.85, bottom=0.15)
        plt.tight_layout()

        # Afficher le graphique
        st.pyplot(fig)


# Ajouter la colonne "Catégorie" pour chaque tâche
def assign_category(task):
    for category, tasks in categories_mapping.items():
        if task in tasks:
            return category
    return "Autre"


# Étape 3 : Résultats
if st.session_state["scores_entered"]:
    st.header("Étape 3 : Résultats")

    age_data = st.session_state["age_data"]
    missing_norms = st.session_state["missing_norms"]


    if "Catégorie" not in age_data.columns:
        categories_mapping = {
            "Langage": [
                "Discrimination Phonologique", "Décision Lexicale Auditive",
                "Mots Outils", "Stock Lexical", "Compréhension Syntaxique"
            ],
            "Mémoire de Travail": [
                "Mémoire de travail verbale endroit empan", "Mémoire de travail verbale endroit brut",
                "Mémoire de travail verbale envers empan", "Mémoire de travail verbale envers brut",
                "Mémoire de travail non verbale endroit empan", "Mémoire de travail non verbale endroit brut",
                "Mémoire de travail non verbale envers empan", "Mémoire de travail non verbale envers brut"
            ],
            "Inhibition": [
                "Inhibition verbale incongruent score",
                "Inhibition verbale incongruent temps",
                "Inhibition non verbale incongruent score",
                "Inhibition non verbale incongruent temps"
            ]
        }
        age_data["Catégorie"] = age_data["Tâche"].apply(assign_category)

    # Afficher le tableau des résultats
    def reorder_columns(dataframe):
    # Liste de colonnes dans l'ordre souhaité
        columns_order = [
            "Tâche",
            "Score Enfant",
            "Z-Score",
            "Moyenne",
            "Ecart-type",
            "Minimum",
            "5e percentile",
            "10e percentile",
            "Q1",
            "Q2 - mediane",
            "Q3",
            "90e percentile",
            "Maximum",
            "Percentile (%)",  # Vous pouvez déplacer cette colonne si nécessaire
        ]
        # Filtrer les colonnes existantes dans le DataFrame selon l'ordre
        reordered_columns = [col for col in columns_order if col in dataframe.columns]
        # Ajouter les colonnes restantes à la fin
        remaining_columns = [col for col in dataframe.columns if col not in reordered_columns]
        return dataframe[reordered_columns + remaining_columns]
    
    st.write("")
    df_to_style = age_data.copy()  # Copie des données originales pour stylisation

    # Réorganiser les colonnes avant tout traitement
    df_to_style = reorder_columns(df_to_style)

    # Formater les nombres en flottants
    def format_floats(value):
        if isinstance(value, float):
            return f"{value:.2f}".rstrip('0').rstrip('.')  # Arrondir à deux décimales et supprimer les zéros inutiles
        return value

    # pandas >= 2.1: DataFrame.map ; pandas plus ancien: DataFrame.applymap
    if hasattr(df_to_style, "map"):
        df_to_style = df_to_style.map(format_floats)
    else:
        df_to_style = df_to_style.applymap(format_floats)
    df_to_style["Percentile (%)"] = pd.to_numeric(df_to_style["Percentile (%)"], errors="coerce")  # Assurez-vous que les percentiles sont numériques

    # Appliquer les styles conditionnels
    def color_percentiles_by_range(value):
        if pd.isna(value):  
            return ''  
        value = float(value)  
        if value <= 3:
            return 'background-color: rgba(212, 70, 70, 0.5); color: black;'  
        elif value <= 15:
            return 'background-color: rgba(245, 167, 47, 0.5); color: black;'  
        elif value <= 85:
            return 'background-color: rgba(96, 205, 114, 0.5); color: black;'  
        elif value <= 97:
            return 'background-color: rgba(141, 223, 155, 0.5); color: black;'  
        elif value <= 100:
            return 'background-color: rgba(174, 222, 182, 0.5); color: black;'  
        return ''  

    def color_task_text_by_category(row):
        category_colors = {
            "Langage": "#3798da",
            "Mémoire de Travail": "#eca113",
            "Inhibition": "#8353da",
            "Autre": "gray"
        }
        category = row["Catégorie"]
        color = category_colors.get(category, "black")
        return [f"color: {color}; font-weight: bold;" if col == "Tâche" else "" for col in row.index]

    # pandas >= 2.1: Styler.map ; pandas plus ancien: Styler.applymap
    if hasattr(df_to_style.style, "map"):
        styled_df = df_to_style.style.map(color_percentiles_by_range, subset=["Percentile (%)"])
    else:
        styled_df = df_to_style.style.applymap(color_percentiles_by_range, subset=["Percentile (%)"])
    styled_df = styled_df.apply(color_task_text_by_category, axis=1)
    
     # Taille colonne
    col_config = {
        styled_df.columns[0]: st.column_config.Column(width=300),  # Première colonne à 300
    }
    col_config.update({col: st.column_config.Column(width=100) for col in styled_df.columns[1:]})  # Le reste à 100


    # Afficher le tableau stylisé dans Streamlit
    st.dataframe(styled_df, hide_index=True, use_container_width=True)

  
    # Sélection des tâches
    st.subheader("Sélectionnez les tâches à afficher dans le graphique")
    calculated_tasks = age_data[~age_data["Z-Score"].isna()]["Tâche"].tolist()
    tasks_by_category = {}
    for category, tasks in categories_mapping.items():
        tasks_in_category = [task for task in tasks if task in calculated_tasks]
        if tasks_in_category:
            tasks_by_category[category] = tasks_in_category

    col1, col2, col3, col4, col5 = st.columns([1, 2, 1, 2, 1])
    with col2:
        if st.button("Tout sélectionner"):
            selected_tasks = calculated_tasks
        else:
            selected_tasks = []

    with col4:
        if st.button("Tout désélectionner"):
            selected_tasks = []

    # Bouton sélection des tâches
    selected_tasks = st.multiselect(
        "Tâches calculées disponibles :", 
        options=calculated_tasks, 
        default=selected_tasks,
        help="Vous pouvez rechercher ou sélectionner des tâches dans la liste."
    )

# Sauvegarde graphique
    def save_styled_excel_to_file(dataframe, file_name="resultats.xlsx"):
        # Réorganiser les colonnes
        dataframe = reorder_columns(dataframe)

        # Initialiser le fichier Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Résultats"

        # Couleurs pour les remplissages conditionnels
        fill_colors = {
            "red": "D44646",
            "orange": "F5A72F",
            "green": "60CD72",
            "light_green": "8DDF9B",
            "blue": "AEDFB6",
        }

        # Couleurs pour les catégories
        category_colors = {
            "Langage": "3798DA",
            "Mémoire de Travail": "ECA113",
            "Inhibition": "8353DA",
            "Autre": "808080",
        }

        # Ajouter les en-têtes
        headers = list(dataframe.columns)
        ws.append(headers)  # Ajout des en-têtes
        header_font = Font(bold=True)
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
            for cell in col:
                cell.font = header_font

        # Ajouter les données ligne par ligne
        for idx, row in dataframe.iterrows():
            ws.append(row.tolist())  # Ajouter la ligne correspondant à l'ordre des colonnes
            excel_row = ws[idx + 2]  # Ligne Excel correspondante (décalée par 1 pour l'en-tête)

            # Appliquer des styles conditionnels
            for col_idx, cell in enumerate(excel_row, start=1):
                # Couleur pour les percentiles
                if headers[col_idx - 1] == "Percentile (%)":
                    try:
                        value = float(cell.value)
                        if value <= 3:
                            fill_color = fill_colors["red"]
                        elif value <= 15:
                            fill_color = fill_colors["orange"]
                        elif value <= 85:
                            fill_color = fill_colors["green"]
                        elif value <= 97:
                            fill_color = fill_colors["light_green"]
                        elif value <= 100:
                            fill_color = fill_colors["blue"]
                        else:
                            fill_color = None

                        if fill_color:
                            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    except (ValueError, TypeError):
                        pass

                # Couleur pour la colonne "Tâche"
                if headers[col_idx - 1] == "Tâche":
                    category = dataframe.loc[idx, "Catégorie"] if "Catégorie" in dataframe.columns else None
                    color = category_colors.get(category, "000000") if category else "000000"
                    cell.font = Font(color=color, bold=True)

        # Sauvegarder le fichier Excel
        try:
            wb.save(file_name)
            st.success(f"Fichier Excel sauvegardé : {file_name}")
        except Exception as e:
            st.error(f"Erreur lors de la sauvegarde du fichier Excel : {e}")


    def save_graph_and_excel(dataframe, selected_tasks, file_name_prefix="resultats"):
        dataframe = reorder_columns(dataframe)
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w") as zf:
            # Graphique
            fig_buffer = io.BytesIO()
            plot_grouped_scores(dataframe, selected_tasks)  # Fonction pour tracer le graphique
            plt.savefig(fig_buffer, format="png", dpi=300, bbox_inches="tight")
            fig_buffer.seek(0)
            zf.writestr(f"{file_name_prefix}_Graphique.png", fig_buffer.read())

            # Excel
            excel_buffer = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Résultats"

            # Ajout des données et style dans Excel
            headers = list(dataframe.columns)
            ws.append(headers)
            header_font = Font(bold=True)
            for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
                for cell in col:
                    cell.font = header_font

            for idx, row in dataframe.iterrows():
                ws.append(row.values.tolist())

            wb.save(excel_buffer)
            excel_buffer.seek(0)
            zf.writestr(f"{file_name_prefix}_Tableau.xlsx", excel_buffer.read())

        buffer.seek(0)
        return buffer


if st.session_state["scores_entered"] and selected_tasks:
    st.subheader("Téléchargez les résultats")
    file_name_prefix = f"{st.session_state['child_id']}_Resultats_Comprendre"
    zip_file = save_graph_and_excel(age_data, selected_tasks, file_name_prefix)
    st.download_button(
        label="📥 Télécharger le tableau des résultats et le graphique (ZIP)",
        data=zip_file,
        file_name=f"{file_name_prefix}.zip",
        mime="application/zip",
    )

# Footer avec citation APA 7
st.markdown(
    """
    <hr style="border:1px solid #eee; margin-top: 50px; margin-bottom: 10px;">
    <div style="text-align: center; font-size: 14px; color: gray;">
    <p><strong>Projet COMPRENDRE</strong> - Pour plus d'informations sur ce projet, consultez le site suivant :<br>
        <a href="https://www.perrone-bertolotti.fr/projet-comprendre" target="_blank">projet-comprendre</a>.
    </p>

    <p>Pour citer le protocole, veuillez utiliser la référence suivante :</p>
        <p style="text-align: center;">
            Perrone-Bertolotti, M., Zoubrinetzky, R., Faure, L., Vaidie, A., Nguyen-Morel, M.-A., Guinet, E., & Gillet-Perret, E. (2023). 
            <em>COMPRENDRE Protocol: A computerized protocol for assessing oral language comprehension and executive functions in French-speaking children aged 5-8</em>. 
            <a href="https://doi.org/10.31234/osf.io/whkcv" target="_blank">https://doi.org/10.31234/osf.io/whkcv</a>
        </p>

    </div>
    """,
    unsafe_allow_html=True
)
